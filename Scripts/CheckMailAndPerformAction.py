import datetime
import win32com.client as win32
import logging
import logging.config
from ExcelOperations import writeToExcel
import traceback
from BatchJobSummary import updateSummaryCounter
from BatchJobSummary import updateJobDetails
from email_handler import send_error_mail
from ReadingProperties import getProperty
import Execution
import openpyxl

logFile= getProperty("MasterConfigurationSection","LoggingConfigFile")
logging.config.fileConfig(logFile)
logger = logging.getLogger('CheckMailAndPerformAction')
configurationFile = getProperty("MasterConfigurationSection","MasterConfigFile")

def checkMail(mail):
	logger.debug('checkMail:: START')
	mailSubject = mail.Subject
	mailBody = mail.Body
	mailHTMLBody = mail.HTMLBody
	mailReceivedTime= mail.ReceivedTime
	applicationName = None
	jobName = None
	contactEmailAddress = None
	#Loop through the config file to get the list of rules
	try:
		if mail.SenderEmailType == "EX":
			if(mail.Sender.GetExchangeUser() is not None):
				mailFromAddress = mail.Sender.GetExchangeUser().PrimarySmtpAddress
			elif(mail.Sender.GetExchangeDistributionList() is not None):
				mailFromAddress = mail.Sender.GetExchangeDistributionList().PrimarySmtpAddress
		else:
			mailFromAddress = mail.SenderEmailAddress 
		mailRecipientsList = mail.Recipients
		
		print("_____________________________________________________")
		print 'Processing new email received from ' + mailFromAddress + ' at ' + format(mailReceivedTime)
		
		workBookConfigurations = openpyxl.load_workbook(filename=configurationFile, read_only=True)
		workSheetConfiguration = workBookConfigurations.get_sheet_by_name("Configuration")
		
		startRow = 2
		endrow = workSheetConfiguration.max_row 
		
		logger.info('From: '+ mailFromAddress)
		logger.info('Subject: '+ mailSubject)
		logger.info('Received Time: '+ format(mailReceivedTime))
		
		logger.debug('Reading the configuration file')
		bConfigurationMatch = False
		for i in range(startRow,endrow):
			applicationName = workSheetConfiguration.cell(row=i, column=1).value
			jobName = workSheetConfiguration.cell(row=i, column=2).value
			
			#Break out of the loop if there is any row with no job name or application name. i.e. consider it as EOF
			if(not (jobName or applicationName)):
				break
			
			#read all the configuration values
			fromEmail =  workSheetConfiguration.cell(row=i, column=3).value 
			toEmail = workSheetConfiguration.cell(row=i, column=4).value 
			subjectLineKeywordsList = workSheetConfiguration.cell(row=i, column=5).value.split('|') if workSheetConfiguration.cell(row=i, column=5).value else []
			
			if not(type(subjectLineKeywordsList) == list):
					subjectLineKeywordsList = [subjectLineKeywordsList]
			mailBodyKeywordsList = workSheetConfiguration.cell(row=i, column=6).value.split('|') if workSheetConfiguration.cell(row=i, column=6).value else []
			
			if not(type(mailBodyKeywordsList) == list):
					mailBodyKeywordsList = [mailBodyKeywordsList]
					
			hasAttachmentFlag = workSheetConfiguration.cell(row=i, column=7).value
			
			attachmentFileNameKeywordsList = workSheetConfiguration.cell(row=i, column=8).value.split('|') if workSheetConfiguration.cell(row=i, column=7).value else []
			
			if not(type(attachmentFileNameKeywordsList) == list):
					attachmentFileNameKeywordsList = [attachmentFileNameKeywordsList]
					
			attachmentType = workSheetConfiguration.cell(row=i, column=9).value
			
			attachmentKeywordsList = workSheetConfiguration.cell(row=i, column=10).value.split('|') if workSheetConfiguration.cell(row=i, column=10).value else []
			
			if not(type(attachmentKeywordsList) == list):
					attachmentKeywordsList = [attachmentKeywordsList]

			successKeywordsList = workSheetConfiguration.cell(row=i, column=11).value.split('|') if workSheetConfiguration.cell(row=i, column=11).value else []
			
			if not(type(successKeywordsList) == list):
					successKeywordsList = [successKeywordsList]
					
			failureKeywordsList = workSheetConfiguration.cell(row=i, column=12).value.split('|') if workSheetConfiguration.cell(row=i, column=12).value else []
			
			if not(type(failureKeywordsList) == list):
					failureKeywordsList = [failureKeywordsList]
					
			successAction1 = workSheetConfiguration.cell(row=i, column=13).value 
			successAction1TargetFileName = workSheetConfiguration.cell(row=i, column=14).value
			successAction1SheetOrCommandLineArgs = workSheetConfiguration.cell(row=i, column=15).value
			successAction2 = workSheetConfiguration.cell(row=i, column=16).value
			successAction2TargetFileName = workSheetConfiguration.cell(row=i, column=17).value
			successAction2SheetOrCommandLineArgs = workSheetConfiguration.cell(row=i, column=18).value
			failureAction1 = workSheetConfiguration.cell(row=i, column=19).value
			failureAction1TargetFileName = workSheetConfiguration.cell(row=i, column=20).value
			failureAction1SheetOrCommandLineArgs = workSheetConfiguration.cell(row=i, column=21).value
			failureAction2 = workSheetConfiguration.cell(row=i, column=22).value
			failureAction2TargetFileName = workSheetConfiguration.cell(row=i, column=23).value
			failureAction2SheetOrCommandLineArgs = workSheetConfiguration.cell(row=i, column=24).value
			contactEmailAddress = workSheetConfiguration.cell(row=i, column=25).value
			
			#Check if the From Email Address matches the configuration value
			if(fromEmail and not(fromEmail.strip().lower() in mailFromAddress.lower())):
				logger.debug('From email address {} does not match the configuration value {}'.format(mailFromAddress,fromEmail))
				continue
			#Check if the TO Email Address matches the configuration value
			if(toEmail):
				bMatchesToEmail = False
				for recipient in mailRecipientsList:
					if(recipient.AddressEntry.GetExchangeUser() is not None):
						emailAddress = recipient.AddressEntry.GetExchangeUser().PrimarySmtpAddress.lower()
					elif (recipient.AddressEntry.GetExchangeDistributionList() is not None):
						emailAddress = recipient.AddressEntry.GetExchangeDistributionList().PrimarySmtpAddress.lower()
					else:
						emailAddress = str(recipient.Address).lower()
						
					if(toEmail.strip().lower() in emailAddress):
						bMatchesToEmail = True
						break
					else:
						logger.debug('To email address {} does not contain the configuration value: {}'.format(emailAddress,toEmail))
					
				if(not bMatchesToEmail):
					logger.debug('To email addresses does not contain the configuration value: {}'.format(toEmail))
					#continue to next configuration in the file
					continue
				
			#Check for Subject Line Keywords List
			if(len(subjectLineKeywordsList) > 0):
				bMatchesSubjectLine = False
				#we are matching any of the keywords here. So if you mention ABCD|END in the keywords, it would look for both the keywords and pass if any one of them is found.
				for keyword in subjectLineKeywordsList:
					if(keyword.strip().lower() in mailSubject.lower()):
						bMatchesSubjectLine = True
						break
				
				if(not bMatchesSubjectLine):
					logger.debug('Subject line of the email does not contain any of the configured keywords: {}'.format(mailSubject))
					#continue to next configuration in the file
					continue
			
			#Check for Mail Body Keywords List
			if(len(mailBodyKeywordsList) > 0):
				bMatchesMailBody = False
				
				#check if the email body has old email chain
				#for HTML or Rich Text Outlook Format
				if "<p class=MsoNormal><b>From:</b>" in mailHTMLBody:
					index = mailHTMLBody.index("<p class=MsoNormal><b>From:</b>")
					mailHTMLBody = mailHTMLBody[:index]
				
				#for plain text/rtf format
				if "-----Original Message-----" in mailHTMLBody:
					index = mailHTMLBody.index("-----Original Message-----")
					mailHTMLBody = mailHTMLBody[:index]

				for keyword in mailBodyKeywordsList:
					if(keyword.strip().lower() in mailHTMLBody.lower()):
						bMatchesMailBody = True
						break
				
				if(not bMatchesMailBody):
					logger.debug('Email body does not contain any of the configured keywords.')
					#continue to next configuration in the file
					continue
					
			#match AttachmentFlag
			if hasAttachmentFlag and hasAttachmentFlag == 'Yes':
				logger.debug('Checking for email attachments.')
				bAttachmentFound = True
				bFileHasKeywords = True
				if mail.Attachments.Count > 0:
					for attachment in mail.Attachments:
						for keyword in attachmentFileNameKeywordsList:
							if keyword.strip().lower() in attachment.FileName.lower():
								bAttachmentFound = True
								break
							else:
								bAttachmentFound = False
							
						#TODO check attachent file for attachmentKeywordsList
						for keyword in attachmentKeywordsList:
							if(attachmentFileHasKeyword(attachment, attachmentType,keyword.strip())):
								bFileHasKeywords = True
								break
							else:
								bFileHasKeywords = False
				
				else: 
					bAttachmentFound = False
					
				if(not (bAttachmentFound or bFileHasKeywords)):
					logger.debug('Email does not contain the attachment as per configuration.')
					#continue to next configuration in the file
					continue
				
			logger.debug('Email has matched all the required criteria per configuration file. Checking for success/failure.')
			bConfigurationMatch = True
		
			bSuccess = False
			for keyword in successKeywordsList:
				#Here we are checking if any of the keywords are matching. in mail body or in subject line 
				if(keyword.strip().lower() in mailBody.lower() or keyword.strip().lower() in mailSubject.lower()):
					logger.debug('Email has matched the success criteria as per configuration')
					bSuccess = True
					break
				else:
					bSuccess = False
			
			if(bSuccess):
				#update the counters in BatchJobSummary
				updateSummaryCounter(True,True)
				#update Job Details
				updateJobDetails(applicationName,jobName,mailReceivedTime,'Success')
				#execute success actions
				if(successAction1):
					logger.debug('Executing Success Action 1')
					performAction(successAction1,successAction1TargetFileName,successAction1SheetOrCommandLineArgs,applicationName,jobName,mailReceivedTime,True,contactEmailAddress,mail)
				if(successAction2):
					logger.debug('Executing Success Action 1')
					performAction(successAction2,successAction2TargetFileName,successAction2SheetOrCommandLineArgs,applicationName,jobName,mailReceivedTime,True,contactEmailAddress,mail)
				break
			else:
				#check for failure
				bFailure = False
				#Here we are checking if any of the keywords are matching. 
				for keyword in failureKeywordsList:
					if(keyword.strip().lower() in mailBody.lower() or keyword.strip().lower() in mailSubject.lower()):
						logger.debug('Email has matched the failure criteria as per configuration')
						bFailure = True
						break
				
				if(bFailure):
					#update the counters in BatchJobSummary
					updateSummaryCounter(True,False)
					#update Job Details
					updateJobDetails(applicationName,jobName,mailReceivedTime,'Failed')
					#execute failureActions
					logger.debug('Executing Failure Actions')
					if(failureAction1):
						logger.debug('Executing Failure Action 1')
						performAction(failureAction1,failureAction1TargetFileName,failureAction1SheetOrCommandLineArgs,applicationName, jobName,mailReceivedTime,False,contactEmailAddress,mail)
					if(failureAction2):
						logger.debug('Executing Failure Action 1')
						performAction(failureAction2,failureAction2TargetFileName,failureAction2SheetOrCommandLineArgs,applicationName, jobName,mailReceivedTime,False,contactEmailAddress,mail)
				break

		workBookConfigurations.close()
		
		if(not bConfigurationMatch):
			logger.debug('The email did not match any of the configurations.')
			print ("The email did not match any of the configurations.")
			#update the counters in BatchJobSummary
			updateSummaryCounter(False,False)
			logger.debug('checkMail:: END')
			print("_____________________________________________________")
			return False
		
		if (bConfigurationMatch and not bSuccess and not bFailure):
			logger.debug('No Success/Failure Keywords Matched')
			print ("No Success/Failure Keywords Matched")
			#update the counters in BatchJobSummary as email did not match the criteria
			updateSummaryCounter(False,False)
			logger.debug('checkMail:: END')
			print("_____________________________________________________")
			return False
		
		print 'Email has been processed'		
		print("_____________________________________________________")
		logger.debug('checkMail:: END')
		return True
	except Exception as err:
		logger.error('checkMail: Error Processing Email')
		logger.error(traceback.format_exc(err))
		send_error_mail( applicationName if applicationName else "System Error", jobName if jobName else "N/A",str(mailReceivedTime),"Error Processing Email.",traceback.format_exc(err),contactEmailAddress if contactEmailAddress else None,None)
		logger.debug('checkMail:: END')
		return False
	
	
	
def performAction(action,targetFileName,sheetorCommandListArgs,applicationName, jobName,mailReceivedTime,bSuccess,contactEmailAddress,mailItem):
	logger.debug('performAction:: START')
	try:
		print 'Processing Action: ' + action
		if action == 'Excel Update':
			logger.debug('Executing Action - Excel Update')
			logger.debug('Excel File Name: ' + targetFileName)
			logger.debug('Sheet Name: ' + sheetorCommandListArgs)
			ret = writeToExcel(targetFileName,sheetorCommandListArgs,applicationName,jobName,mailReceivedTime,bSuccess,contactEmailAddress)
		elif action == 'Execute BAT File':
			logger.debug('Executing Action - Execute BAT File')
			logger.debug('File Name: ' + targetFileName)
			logger.debug('Command Line Args: ' + sheetorCommandListArgs)
			ret = Execution.executeAction('BAT',targetFileName,sheetorCommandListArgs,applicationName,jobName,mailReceivedTime,contactEmailAddress,mailItem)
		elif action == 'Execute Python Script':
			logger.debug('Executing Action - Execute Python Script')
			logger.debug('File Name: ' + targetFileName)
			logger.debug('Command Line Args: ' + sheetorCommandListArgs)
			ret = Execution.executeAction('PY',targetFileName,sheetorCommandListArgs,applicationName,jobName,mailReceivedTime,contactEmailAddress,mailItem)
		elif action == 'Execute EXE':
			logger.debug('Executing Action - Execute EXE')
			logger.debug('File Name: ' + targetFileName)
			logger.debug('Command Line Args: ' + sheetorCommandListArgs)
			ret = Execution.executeAction('EXE',targetFileName,sheetorCommandListArgs,applicationName,jobName,mailReceivedTime,contactEmailAddress,mailItem)
		elif action == 'Execute JAR':
			logger.debug('Executing Action - Execute JAR')
			logger.debug('File Name: ' + targetFileName)
			logger.debug('Command Line Args: ' + sheetorCommandListArgs)
			ret = Execution.executeAction('JAR',targetFileName,sheetorCommandListArgs,applicationName,jobName,mailReceivedTime,contactEmailAddress,mailItem)
		logger.debug('performAction:: END')
	except Exception as err:
		logger.error('performAction: Error Performing Action')
		logger.error(traceback.format_exc(err))
		print("Error Performing Action")
	
def attachmentFileHasKeyword(attachent,attachmentType,keyword):
	logger.debug('attachmentFileHasKeyword:: START')
	return True #to be further implemented
	logger.debug('attachmentFileHasKeyword:: END')