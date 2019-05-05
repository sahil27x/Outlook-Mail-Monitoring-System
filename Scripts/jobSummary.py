import openpyxl
import datetime
import time
from os.path import exists
import traceback
import logging
import logging.config
import numpy as np
import matplotlib.pyplot as plt
from ReadingProperties import getProperty
from BatchJobDetails import JobDetails, JobSta1tus
import shutil
from openpyxl.styles import Border, Side, Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

logFile= getProperty("MasterConfigurationSection","LoggingConfigFile")
logging.config.fileConfig(logFile)
logger = logging.getLogger('BatchJobSummary')
fileName=getProperty('BatchJobSummary','fileName')
txtFileName=getProperty('BatchJobSummary','txtFileName')
jobDetailsFileName=getProperty('BatchJobDetails','jobDetailsFileName')


def updateSummaryCounter(matchingEmailFlag,successEmailFlag):
	try:
		logger.info('updateSummaryCounter:: START')
		todaysDate = datetime.datetime.now().strftime('%m/%d/%Y')
		totalemails=[]
		matchingemails=[]
		successemails=[]
		failedemails=[]
		dates=[]
		# variables for summary details
		total_emails=0
		matched_emails=0
		success_emails=0
		failed_emails=0

		logger.info('Batch Job Status File: {}'.format(fileName))
		thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
		alignment=Alignment(horizontal='general', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
		if exists(fileName):
			logger.info('Batch Job Status File Found')
			
			wb = openpyxl.load_workbook(fileName) 
			ws = wb.get_sheet_by_name("BatchJobSummary")
			
			row=ws.max_row

			try:
				excelDate = time.strptime(str(ws.cell(row=row, column=1).value), '%m/%d/%Y')
			except Exception as e:
				excelDate = time.strptime(str(ws.cell(row=row, column=1).value), '%Y-%m-%d %H:%M:%S')
			excelDate = (time.strftime('%m/%d/%Y', excelDate))
			
			if todaysDate != excelDate:
				row =row+1
				ws.cell(row=row,column=1).value = todaysDate

			#Update Total Emails Counter
			if ws.cell(row=row,column=2).value is None:
				ws.cell(row=row,column=2).value = 1
			else:
				ws.cell(row=row,column=2).value = ws.cell(row=row,column=2).value+1
			
			#Update Matching Emails Counter
			if matchingEmailFlag == True:
				if ws.cell(row=row,column=3).value is None:
					ws.cell(row=row,column=3).value = 1
				else:
					ws.cell(row=row,column=3).value = ws.cell(row=row,column=3).value + 1

				#Update Success Emails Counter
				if successEmailFlag == True:
					if ws.cell(row=row,column=4).value is None:
						ws.cell(row=row,column=4).value =1
					else:
						ws.cell(row=row,column=4).value = ws.cell(row=row,column=4).value + 1

				#update Failure Emails Counter
				else:
					if ws.cell(row=row,column=5).value is None:
						ws.cell(row=row,column=5).value = 1
					else:
						ws.cell(row=row,column=5).value = ws.cell(row=row,column=5).value + 1
			
			for i in range(1,6):
				ws.cell(row=row, column=i).border = thin_border
				ws.column_dimensions[get_column_letter(i)].width = 25
				rows = ws.row_dimensions[row]
				rows.alignment=alignment
			
			if row > 6:
				startRow = row - 4 # we will generate graph only for last 5 days
			else:
				startRow = 2
			
			for i in range(startRow,row+1):
				if ws.cell(row=i,column=1).value is not None:
					try:
						chartDate = time.strptime(str(ws.cell(row=i,column=1).value), '%m/%d/%Y')
					except Exception as e:
						chartDate = time.strptime(str(ws.cell(row=i,column=1).value), '%Y-%m-%d %H:%M:%S')
					chartDate = (time.strftime('%m/%d/%Y', chartDate))
					dates.append(chartDate)
				if ws.cell(row=i,column=2).value is not None:
					totalemails.append(int(ws.cell(row=i,column=2).value))
					total_emails=int(ws.cell(row=i,column=2).value)
				else:
					totalemails.append(0) 
				if ws.cell(row=i,column=3).value is not None:
					matchingemails.append(int(ws.cell(row=i,column=3).value))
					matched_emails = int(ws.cell(row=i,column=3).value)
				else:
					matchingemails.append(0) 
				if ws.cell(row=i,column=4).value is not None:
					successemails.append(int(ws.cell(row=i,column=4).value))
					success_emails = int(ws.cell(row=i,column=4).value)
				else:
					successemails.append(0) 
				if ws.cell(row=i,column=5).value is not None:
					failedemails.append(int(ws.cell(row=i,column=5).value))
					failed_emails = int(ws.cell(row=i,column=5).value)
				else:
					failedemails.append(0) 
			
			wb.save(fileName)
			wb.close()
			logger.info('Details Updated in Batch Job Status File')
		else:
			logger.info('Batch Job Status File Not Found, Creating New File')
			wb = openpyxl.Workbook()
			ws = wb.active
			ws.title = "BatchJobSummary"
			
			ws.cell(row=1,column=1).value = "Date"
			ws.cell(row=1,column=2).value = "No. of Processed Emails"
			ws.cell(row=1,column=3).value = "No. of Matching Emails"
			ws.cell(row=1,column=4).value = "No. of Success Emails"
			ws.cell(row=1,column=5).value = "No. of Error Emails"
			
			for i in range(1,6):
				ws.cell(row=1, column=i).border = thin_border
				ws.column_dimensions[get_column_letter(i)].width = 25
				rows = ws.row_dimensions[1]
				rows.alignment=alignment
			
			row=2
			ws.cell(row=row,column=1).value = todaysDate
			ws.cell(row=row,column=2).value = 0
			ws.cell(row=row,column=3).value = 0
			ws.cell(row=row,column=4).value = 0
			ws.cell(row=row,column=5).value = 0


			#Update Total Emails Counter
			if ws.cell(row=row,column=2).value is None:
				ws.cell(row=row,column=2).value = 1
			else:
				ws.cell(row=row,column=2).value = ws.cell(row=row,column=2).value + 1
			
			#Update Matching Emails Counter
			if matchingEmailFlag == True:
				if ws.cell(row=row,column=3).value is None:
					ws.cell(row=row,column=3).value = 1
				else:
					ws.cell(row=row,column=3).value = ws.cell(row=row,column=3).value + 1

				#Update Success Emails Counter
				if successEmailFlag == True:
					if ws.cell(row=row,column=4).value is None:
						ws.cell(row=row,column=4).value =1
					else:
						ws.cell(row=row,column=4).value = ws.cell(row=row,column=4).value + 1

				#update Failure Emails Counter
				else:
					if ws.cell(row=row,column=5).value is None:
						ws.cell(row=row,column=5).value = 1
					else:
						ws.cell(row=row,column=5).value = ws.cell(row=row,column=5).value + 1
			
			for i in range(1,6):
				ws.cell(row=row, column=i).border = thin_border
				ws.column_dimensions[get_column_letter(i)].width = 25
				rows = ws.row_dimensions[row]
				rows.alignment=alignment
			
			if row > 6:
				startRow = row - 4 # we will generate graph only for last 5 days
			else:
				startRow = 2
			
			for i in range(startRow,row+1):
				if ws.cell(row=i,column=1).value is not None:
					try:
						chartDate = time.strptime(str(ws.cell(row=i,column=1).value), '%m/%d/%Y')
					except Exception as e:
						chartDate = time.strptime(str(ws.cell(row=i,column=1).value), '%Y-%m-%d %H:%M:%S')
					chartDate = (time.strftime('%m/%d/%Y', chartDate))
					dates.append(chartDate)
				if ws.cell(row=i,column=2).value is not None:
					totalemails.append(int(ws.cell(row=i,column=2).value))
					total_emails=int(ws.cell(row=i,column=2).value)
				else:
					totalemails.append(0) 
				if ws.cell(row=i,column=3).value is not None:
					matchingemails.append(int(ws.cell(row=i,column=3).value))
					matched_emails = int(ws.cell(row=i,column=3).value)
				else:
					matchingemails.append(0) 
				if ws.cell(row=i,column=4).value is not None:
					successemails.append(int(ws.cell(row=i,column=4).value))
					success_emails = int(ws.cell(row=i,column=4).value)
				else:
					successemails.append(0) 
				if ws.cell(row=i,column=5).value is not None:
					failedemails.append(int(ws.cell(row=i,column=5).value))
					failed_emails = int(ws.cell(row=i,column=5).value)
				else:
					failedemails.append(0) 

			wb.save(fileName)
			wb.close()
			logger.info('Details Updated in Batch Job Status File')
			logger.info('updateSummaryCounter:: END')
			logger.info('Creation Of Summary Graph')
		try:
			logger.info('Writing to file: '.format(txtFileName))
			fo = open(txtFileName, "w+") # Creation of text file for batch job summary
			#print(todaysDate+','+str(total_emails)+','+str(matched_emails)+','+str(success_emails)+','+str(failed_emails))
			fo.write(todaysDate+','+str(total_emails)+','+str(matched_emails)+','+str(success_emails)+','+str(failed_emails))
			fo.close()
		except Exception as err:
			logger.error('Error Writing  to BatchSummary.txt:'+ traceback.format_exc(err))
			print traceback.format_exc(err)
		generateSummaryGraph(totalemails, matchingemails, successemails,failedemails,dates)
	except Exception as err:
		logger.error('writeToExcel: Error Writing to Target Excel File')
		logger.error(traceback.format_exc(err))
		#sendEmail(err)
		print("Error Writing to Batch Job Status File")
	
	logger.debug('writeToTextFile:: END')


#methods for creating summary graph
def generateSummaryGraph(totalemails, matchingemails, successemails,failedemails,dates):
	try:
		logger.info('PlotGraph: START')
		N = len(totalemails)
		ind = np.arange(N)  # the x locations for the groups(total number od days for which data is received)
		width = 0.20	   # the width of the bars
		fig = plt.figure()
		ax = fig.add_subplot(111)

		emailsTotal = totalemails #Total emails recieved
		rects0 = ax.bar(ind, emailsTotal, width, color='b') #Plot the bar chart in blue color

		emailsMatched = matchingemails # Total matching emails
		rects1 = ax.bar(ind+width, emailsMatched, width, color='c') #Plot the bar chart in cyan color

		emailsSuccess = successemails # Total success emails
		rects2 = ax.bar(ind+width*2, emailsSuccess, width, color='g') #Plot the bar chart in green color

		emailsFailed = failedemails # Total failed emails
		rects3 = ax.bar(ind+width*3, emailsFailed, width, color='r') #Plot the bar chart in red color

		ax.set_ylabel('Number of Emails') # Label for Y axis
		ax.set_xlabel('Date') # Label for X axis
		ax.set_xticks(ind+width)
		labels=dates
		ax.set_xticklabels(labels)
		ax.legend( (rects0[0],rects1[0], rects2[0], rects3[0]), ('No. of Processed Emails','No. of Matching Emails', 'No. of Jobs Success', 'No. of Jobs Failed') )#lengend for graph
		def autolabel(rects):
			for rect in rects:
				h = rect.get_height()
				ax.text(rect.get_x()+rect.get_width()/2., 1.002*h, '%d'%int(h),
				ha='center', va='bottom')
		autolabel(rects0)
		autolabel(rects1)
		autolabel(rects2)
		autolabel(rects3)
		
		# Saving the graph to directory specified in config file
		plt.savefig(getProperty('BatchJobSummary','summaryGraphPath') + getProperty('BatchJobSummary','summaryGraphFileName')) 
		del fig
		plt.close()
	except Exception as err:
		logger.error('PlotGraph: Error In Creating Graph')
		logger.error(traceback.format_exc(err))
		print traceback.format_exc(err)
	logger.info('PlotGraph: END')

#method to get summary details
def getSummaryDetails():
	try:
		logger.info('getSummaryDetails:START')
		logger.info('Reading from file: {}'.format(txtFileName))
		if exists(txtFileName):
			fo = open(txtFileName, "r")
			summaryDetails = fo.read(); #Read the details
			fo.close()
			return summaryDetails
		else:
			return None # Returns none if file not found
	except Exception as err:
		logger.error('Error in reading summary file: {}'+ traceback.format_exc(err))
		#print err
		return None
	logger.info('getSummaryDetails:END')
	
#method to get job details
def getJobDetails():
	try:
		logger.info('getJobDetails:START')
		logger.info('Reading from file: {}'.format(txtFileName))
		currentCount=getCurrentCount()		
		return  currentCount
	except Exception as err:
		return None
	logger.info('getJobDetails:END')

#method for updating job details to test file and Return the expected count , cuurent count in a dictionary variable
def updateJobDetails(applicationName,programName,receiptTime,batchStatus):
	try:
		logger.info('updateJobDetails:: START')
		datetime.datetime.now().strftime('%m/%d/%y')
		currDate=datetime.datetime.now().strftime('%m%d%Y')# variable appends with file name for creating new file with current date
		fileName=jobDetailsFileName+currDate+".txt"
		logger.info('Writing  to '+fileName)
		new_file= jobDetailsFileName + 'new_file.txt'
		
		checked = False		
		if not exists(fileName): # Create a new file if file does not exists
			try:
				logger.info('Create New file :'+fileName)
				fo = open(fileName,'w+')
				fo.write('Application Name|JobName|Status|Time\n') #First header line to know what all attributes adding to file
				fo.write(applicationName+'|'+programName+'|'+batchStatus+'|'+str(receiptTime)+'\n') #Writing the details to the file created
			finally:
				fo.close()
				
		else :
			#updating the Status and time with new values in text file
			try:
				input_file = open(fileName,'r')
				output_file = open(new_file,'w')
				
				for line in input_file:
					line_array=line.split("|")
					if applicationName == line_array[0] and programName == line_array[1]:
						checked = True
						line3 = line_array[3].strip('\n')
						newLine = applicationName+'|'+programName+'|'+line_array[2]+'&'+batchStatus+'|'+line3+'&'+str(receiptTime)+'\n'
						output_file.write(newLine)
					else:
						output_file.write(line)	
				output_file.close()
				if checked == False:
					output_file=open(new_file,'a')
					output_file.write(applicationName+'|'+programName+'|'+batchStatus+'|'+str(receiptTime)+'\n')
			
			finally:
				input_file.close()
				output_file.close()
				shutil.move(new_file,fileName)
	except Exception as ex:
		logger.error('Error Writing details to text file')
		logger.error(traceback.format_exc(ex))
		print traceback.format_exc(ex)
		
	logger.debug('updateJobDetails:: END')

def getCurrentCount():
	try:
		logger.info('getCurrentCount():: START')
		jobScheduleFileName=getProperty('MasterConfigurationSection','jobScheduleFileName')
		if (not jobScheduleFileName or not exists(jobScheduleFileName)):
			logger.info('getCurrentCount():: Batch Job Schedule File Not Found')
			logger.info('getCurrentCount():: END')
			return None
		
		logger.info('getCurrentCount():: Open the job schedule file ')
		workbook = openpyxl.load_workbook(filename=jobScheduleFileName, read_only=True)
		worksheet = workbook.get_sheet_by_name("Sheet1")
		today = datetime.date.today().weekday()
		day_of_month = str(datetime.date.today().day)
		endrow= worksheet.max_row
		actualJobData=[]
		for i in range(3,endrow):
			if (today == 0 and worksheet.cell(row=i, column=3).value is not None and worksheet.cell(row=i, column=3).value == 'Y' and worksheet.cell(row=i,column=4).value is not None): #Monday
				objBatchJobDetails = generateCurrentCount(str(worksheet.cell(row=i, column=1).value),str(worksheet.cell(row=i, column=2).value),str(worksheet.cell(row=i, column=4).value))
				actualJobData.append(objBatchJobDetails)
			elif(today == 1 and worksheet.cell(row=i, column=5).value is not None and worksheet.cell(row=i, column=5).value == 'Y'and worksheet.cell(row=i,column=6).value is not None):#Tuesday
				objBatchJobDetails = generateCurrentCount(str(worksheet.cell(row=i, column=1).value),str(worksheet.cell(row=i, column=2).value),str(worksheet.cell(row=i, column=6).value))
				actualJobData.append(objBatchJobDetails)
			elif(today == 2 and worksheet.cell(row=i, column=7).value is not None and worksheet.cell(row=i, column=7).value=='Y' and worksheet.cell(row=i,column=8).value is not None):#Wed
				objBatchJobDetails = generateCurrentCount(str(worksheet.cell(row=i, column=1).value),str(worksheet.cell(row=i, column=2).value),str(worksheet.cell(row=i, column=8).value))
				actualJobData.append(objBatchJobDetails)
			elif(today == 3 and worksheet.cell(row=i, column=9).value is not None and worksheet.cell(row=i, column=9).value =='Y' and worksheet.cell(row=i,column=10).value is not None):#THU
				objBatchJobDetails = generateCurrentCount(str(worksheet.cell(row=i, column=1).value),str(worksheet.cell(row=i, column=2).value),str(worksheet.cell(row=i, column=10).value))
				actualJobData.append(objBatchJobDetails)
			elif(today == 4 and worksheet.cell(row=i,column=11).value is not None and worksheet.cell(row=i,column=11).value == 'Y' and worksheet.cell(row=i,column=12).value is not None ):#FRI
				objBatchJobDetails = generateCurrentCount(str(worksheet.cell(row=i,column=1).value),str(worksheet.cell(row=i,column=2).value),str(worksheet.cell(row=i,column=12).value))
				actualJobData.append(objBatchJobDetails)
			elif(today == 5 and worksheet.cell(row=i, column=13).value is not None and worksheet.cell(row=i, column=13).value == 'Y' and worksheet.cell(row=i, column=14).value is not None):#SAT
				objBatchJobDetails = generateCurrentCount(str(worksheet.cell(row=i, column=1).value),str(worksheet.cell(row=i, column=2).value),str(worksheet.cell(row=i, column=14).value))
				actualJobData.append(objBatchJobDetails)
			elif(today == 6 and worksheet.cell(row=i, column=15).value is not None and worksheet.cell(row=i, column=15).value == 'Y' and worksheet.cell(row=i, column=16).value is not None):#SUN
				objBatchJobDetails = generateCurrentCount(str(worksheet.cell(row=i, column=1).value),str(worksheet.cell(row=i, column=2).value),str(worksheet.cell(row=i, column=16).value))
				actualJobData.append(objBatchJobDetails)
			elif(worksheet.cell(row=i,column=17).value is not None and day_of_month == str(worksheet.cell(row=i,column=17).value) and worksheet.cell(row=i,column=18).value is not None): #Day of Month Calucation
				objBatchJobDetails = generateCurrentCount(str(worksheet.cell(row=i, column=1).value),str(worksheet.cell(row=i, column=2).value),str(worksheet.cell(row=i, column=18).value))
				actualJobData.append(objBatchJobDetails)
				
		workbook.close()
		return actualJobData
	except Exception as err:
		logger.error('Error  in processing current status details: {}'+ traceback.format_exc(err))
		return None
	logger.debug('getCurrentCount():: END')
	
def generateCurrentCount(appName,jobName,expectedTime):
	try:
		currDate=datetime.datetime.now().strftime('%m%d%Y')# variable appends with file name for creating new file with current date
		logger.info('getCurrentCount():: Open the job details text file in Read mode')
		fileName=jobDetailsFileName+currDate+".txt"
		#pdb.set_trace()
		expectedTimeArray=[] # Array for expected
		actualRunTimeAndStatus=[]
		#statusArray=[]
		timeArray = expectedTime.split("|")
		
		for time_item in timeArray:
			if time_item is not None and time_item is not "":
				try:
					et = time.strptime(time_item,'%I:%M:%S %p') #this format works with openpyXL as time is considered as string
					#et=time_item.strftime('%H:%M:%S') #dileep
					time_expected = datetime.time(et.tm_hour,et.tm_min,et.tm_sec) 
					#expected_time_am_format=datetime.datetime.strptime(str(time_expected),'%H:%M:%S').strftime('%I:%M %p')#print the time in AM/PM format
					expectedTimeArray.append(time_expected)
				except Exception as err:
					logger.error('Invalid Time Format: {} : {}' + time_item + traceback.format_exc(err))
				#expectedTimeArray.append(time_item)
		
		if len(expectedTimeArray) > 0:
			current_time = datetime.time(int(datetime.datetime.now().strftime('%H')),int(datetime.datetime.now().strftime('%M')))
			#et = time.strptime(expectedTimeArray[0],'%I:%M %p') #this format works with openpyXL as time is considered as string
			#time_expected = datetime.time(et.tm_hour,et.tm_min,et.tm_sec) 
			time_expected = expectedTimeArray[0]
			if current_time > time_expected:
				currentJob = JobDetails(appName,jobName,expectedTimeArray,[],"F")
			else:
				currentJob = JobDetails(appName,jobName,expectedTimeArray,[],"T")
		else:
			logger.info("No Expected Run Time Found")
			return None
				
		if not exists(fileName):
			#No Jobs have run so far
			return currentJob
			
		f = open(fileName,'r')
		lines=f.readlines()
		f.close()
		count = []
		for line in lines:
			line_array = line.split("|")
	
			logger.info('getCurrentCount():: Adding the current times and statuses to array')
			if(appName == line_array[0] and jobName == line_array[1]):
				processed_time_line=line_array[3].split("&")
				status_time_line=line_array[2].split("&")
				
				if len(processed_time_line) != len(status_time_line):
					log.error("Invalid Values for the current job in the Job Details File {}. Appliaction Name: {} and Job Name: {}" + fileName + appName + jobName)
					return currentJob
				
				for i in range(0,len(processed_time_line)):
					processed_time=processed_time_line[i].strip()
					status = status_time_line[i].strip()
					processed_time=datetime.datetime.strptime(processed_time,'%m/%d/%y %H:%M:%S')
					processed_time_am_format=processed_time.strftime('%I:%M %p')
					
					jobStatus = JobStatus(status,processed_time)
					actualRunTimeAndStatus.append(jobStatus)
				
				break
		
		currentJob.actualRunTimeAndStatus = actualRunTimeAndStatus
		
		#check the jobs rans on time or not
		
		#In case of the job has not run yet
		if len(actualRunTimeAndStatus) == 0:
			 currentJob.actualRunTimeAndStatus = []
		#If job is expected to run only once, compare expected time with last run instance and if the job ran successfully
		elif len(expectedTimeArray) == 1:
			actualRunDateTime = actualRunTimeAndStatus[(len(actualRunTimeAndStatus) - 1)].actualRunTime
			actualRunTime = datetime.time(int(actualRunDateTime.strftime('%H')),int(actualRunDateTime.strftime('%M')))
			status = actualRunTimeAndStatus[(len(actualRunTimeAndStatus) - 1)].jobStatus
			
			if expectedTimeArray[0] >= actualRunTime and status.lower() == "success":
				currentJob.slaFlag = "T"
			else:
				currentJob.slaFlag = "F"
				
		elif len(actualRunTimeAndStatus) != len(expectedTimeArray):
			currentJob.slaFlag = "F"
			
		elif len(actualRunTimeAndStatus) == len(expectedTimeArray):
			actualRunDateTime = actualRunTimeAndStatus[(len(actualRunTimeAndStatus) - 1)].actualRunTime
			actualRunTime = datetime.time(int(actualRunDateTime.strftime('%H')),int(actualRunDateTime.strftime('%M')))
			status = actualRunTimeAndStatus[(len(actualRunTimeAndStatus) - 1)].jobStatus
			if expectedTimeArray[(len(actualRunTimeAndStatus) - 1)] >= actualRunTime and status.lower() == "success":
				currentJob.slaFlag = "T"
			else:
				currentJob.slaFlag = "F"

	except Exception as err:
		logger.error('Error in processing the status summaries: {}'+ traceback.format_exc(err))
		currentJob = None
	finally:
		logger.debug('generateCurrentCount():: END')
		return currentJob
	

	
def checkDependency(dependantJobName,fileName,processed_time):
	try:
		logger.info('checkDependency():: Opens the job details text file in Read mode')
		f = open(fileName,'r')
		lines=f.readlines()
		f.close()
		checkFlag=False
		for line in lines:
			line_array = line.split("|")
			logger.info('checkDependency():: Checking the job details text file for dependant file name'+dependantJobName+'===='+line_array[1])
			if(dependantJobName == line_array[1] and dependantJobName is not None ):
				parent_time=line_array[3].strip()
				if parent_time != "Time":
					parent_time=datetime.datetime.strptime(parent_time,'%m/%d/%y %H:%M:%S')
					parent_time = datetime.time(int(parent_time.strftime('%H')),int(parent_time.strftime('%M')))
					logger.info('checkDependency()::'+parent_time+"====="+processed_time)
				if(parent_time < processed_time and line_array[2].lower()=='success' ):
					checkFlag = True
				break
		return checkFlag
	except Exception as err:
		logger.error('Error in processing the status summaries: {}'+ traceback.format_exc(err))
		return False
	logger.debug('checkDependency():: END')